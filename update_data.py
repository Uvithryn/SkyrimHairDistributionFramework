#!/usr/bin/env python3
"""One-step data update for the SPID File Generator.

Run this after adding, renaming, or deleting preview screenshots, or after
editing the spreadsheet:

    python update_data.py

It does the whole pipeline in order:
  1. Tidies image and folder names, then rebuilds previews.json from Images/.
  2. Rebuilds Data/*.json from the spreadsheet, validating as it goes.
  3. Refreshes the sample data embedded in index.html (optional).

Nothing is passed on the command line -- edit the settings block below. Paths
are resolved relative to THIS file, so keep it beside index.html and run it
from wherever you like.

Two reports are written next to this script; read them before committing:
  preview_report.txt -- renames, name collisions, folder/filename mismatches
  export_report.txt  -- missing previews, orphaned images, data validation
"""

# ============================ SETTINGS (edit me) ============================
XLSX          = "Skyrim Hairstyle Distribution Framework.xlsx"  # the spreadsheet
IMAGES_ROOT   = "Images"        # folder holding the per-mod screenshot folders
MANIFEST      = "previews.json" # image manifest, rebuilt by step 1
DATA_DIR      = "Data"          # JSON output folder -- must match DATA_DIR in
                                # index.html EXACTLY (case matters once hosted)

# Optional: adds a ready-made "url" to every manifest record. Leave as None
# unless you want absolute URLs baked into previews.json.
IMAGE_URL_BASE = None

# Reports are written to a temp folder while running and only saved next to this
# script if you ask for them at the end, so the repo stays clean.
PREVIEW_REPORT = "preview_report.txt"
EXPORT_REPORT  = "export_report.txt"

# Step 3 rewrites index.html in place. Set False to leave that file alone.
REBUILD_EMBEDDED_SAMPLE = True
INDEX_HTML              = "index.html"
SAMPLE_NPCS_PER_GROUP   = 5   # NPCs kept per race/gender combination
SAMPLE_PREVIEWS_PER_MOD = 6   # preview records kept per mod/race/gender
# ===========================================================================

import json
import os
import re
import shutil
import sys
import tempfile
from collections import Counter, defaultdict
from datetime import datetime, timezone
from pathlib import Path
from types import SimpleNamespace

try:
    import openpyxl
except ImportError:
    print("This script needs openpyxl.  Install it with:  pip install openpyxl", file=sys.stderr)
    raise SystemExit(1)

HERE = Path(__file__).resolve().parent
sys.dont_write_bytecode = True


def rel(name):
    """Resolve a setting against this script's folder."""
    p = Path(name)
    return p if p.is_absolute() else HERE / p


# ===========================================================================
# PART 1 -- image sanitizing and manifest building
# (was sanitize_and_manifest.py)
# ===========================================================================
UNSAFE_RE = re.compile(r"[^A-Za-z0-9._-]")
WHITESPACE_RE = re.compile(r"\s+")

# The race types and genders your framework uses. Folder names are validated
# against these so capture-script typos surface instead of silently 404ing.
KNOWN_RACES = {"Argonian", "Elf", "Human", "HumanChild", "Khajiit", "Orc"}
KNOWN_GENDERS = {"Male", "Female"}

IMAGE_EXTS = {".jpg", ".jpeg", ".png", ".webp"}
JUNK_NAMES = {"thumbs.db", "desktop.ini", ".ds_store"}


def sanitize_name(name: str) -> str:
    """Make a single path component safe for a URL. Idempotent."""
    name = WHITESPACE_RE.sub("_", name)
    name = UNSAFE_RE.sub("", name)
    return name


def parse_stem(stem: str):
    """Split a filename stem into (hairstyle_key, file_race, file_gender).

    Strips a trailing _<Gender> then a trailing _<Race> only when they match the
    known sets, so a hairstyle name that happens to contain an underscore is not
    chewed up. Returns file_race/file_gender = None when no known suffix found.
    """
    parts = stem.split("_")
    file_gender = None
    file_race = None
    if parts and parts[-1] in KNOWN_GENDERS:
        file_gender = parts[-1]
        parts = parts[:-1]
    if parts and parts[-1] in KNOWN_RACES:
        file_race = parts[-1]
        parts = parts[:-1]
    key = "_".join(parts) if parts else stem
    return key, file_race, file_gender


def plan_renames(root: Path):
    """Return a list of (old_path, new_path) and a list of collision errors.

    Renames are ordered so they can be applied safely: files first, then
    directories deepest-first, so no parent is renamed before its children.
    """
    files, dirs = [], []
    for dirpath, dirnames, filenames in os.walk(root):
        for fn in filenames:
            files.append(Path(dirpath) / fn)
        for dn in dirnames:
            dirs.append(Path(dirpath) / dn)

    ordered = files + sorted(dirs, key=lambda p: len(p.parts), reverse=True)

    renames, collisions = [], []
    for p in ordered:
        new_name = sanitize_name(p.name)
        if new_name == p.name:
            continue
        target = p.with_name(new_name)
        if target.exists() and target.resolve() != p.resolve():
            collisions.append((p, target))
        else:
            renames.append((p, target))
    return renames, collisions


def apply_renames(renames):
    done = 0
    for old, new in renames:
        old.rename(new)
        done += 1
    return done


def build_manifest(root: Path, url_base: str | None):
    """Walk the tree and build manifest records + warnings.

    Path components are passed through sanitize_name() in-memory so the result
    matches the post-sanitize on-disk layout regardless of whether --apply ran.
    """
    records = []
    warnings = []
    skipped = []
    mods, races, genders = Counter(), Counter(), Counter()

    for dirpath, dirnames, filenames in os.walk(root):
        for fn in filenames:
            full = Path(dirpath) / fn
            if fn.lower() in JUNK_NAMES or fn.startswith("."):
                skipped.append(("junk", str(full.relative_to(root))))
                continue
            ext = os.path.splitext(fn)[1].lower()
            if ext not in IMAGE_EXTS:
                skipped.append(("non-image", str(full.relative_to(root))))
                continue

            rel_raw = full.relative_to(root)
            # Sanitize every component so the manifest matches final names.
            comps = [sanitize_name(c) for c in rel_raw.parts]
            rel = "/".join(comps)

            if len(comps) != 4:
                warnings.append(
                    f"UNEXPECTED DEPTH ({len(comps)} levels, expected Mod/Race/Gender/file): {rel}"
                )
                # Still record it so nothing is silently lost.
                mod = comps[0] if comps else "?"
                race = gender = "?"
                stem = os.path.splitext(comps[-1])[0]
                key, _, _ = parse_stem(stem)
                rec = {"mod": mod, "race": race, "gender": gender,
                       "hairstyle": key, "file": rel, "mismatch": True}
                if url_base:
                    rec["url"] = url_base.rstrip("/") + "/" + rel
                records.append(rec)
                continue

            mod, race, gender, filename = comps
            stem = os.path.splitext(filename)[0]
            key, frace, fgender = parse_stem(stem)

            if race not in KNOWN_RACES:
                warnings.append(f"UNKNOWN RACE FOLDER '{race}': {rel}")
            if gender not in KNOWN_GENDERS:
                warnings.append(f"UNKNOWN GENDER FOLDER '{gender}': {rel}")

            mismatch = (frace is not None and frace != race) or \
                       (fgender is not None and fgender != gender)
            if mismatch:
                warnings.append(
                    f"MISMATCH: file says ({frace or '?'}/{fgender or '?'}) "
                    f"but folder is ({race}/{gender}) -> {rel}"
                )

            mods[mod] += 1
            races[race] += 1
            genders[gender] += 1

            rec = {"mod": mod, "race": race, "gender": gender,
                   "hairstyle": key, "file": rel, "mismatch": mismatch}
            if url_base:
                rec["url"] = url_base.rstrip("/") + "/" + rel
            records.append(rec)

    records.sort(key=lambda r: r["file"])
    stats = {
        "total_images": len(records),
        "by_mod": dict(sorted(mods.items())),
        "by_race": dict(sorted(races.items())),
        "by_gender": dict(sorted(genders.items())),
        "mismatches": sum(1 for r in records if r["mismatch"]),
        "skipped": len(skipped),
    }
    return records, warnings, skipped, stats


def write_report(path: Path, renames, collisions, records, warnings, skipped, stats, applied):
    lines = []
    lines.append("SMP Hairstyle Preview -- sanitize & manifest report")
    lines.append("generated: " + datetime.now(timezone.utc).isoformat(timespec="seconds"))
    lines.append("mode: " + ("APPLIED (files renamed, manifest written)" if applied else "DRY RUN (nothing written to disk except this preview)"))
    lines.append("")
    lines.append("SUMMARY")
    lines.append(f"  images in manifest : {stats['total_images']}")
    lines.append(f"  mismatched files   : {stats['mismatches']}")
    lines.append(f"  renames needed     : {len(renames)}")
    lines.append(f"  name collisions    : {len(collisions)}")
    lines.append(f"  skipped (junk/non-image): {stats['skipped']}")
    lines.append("")
    lines.append("  by mod:")
    for k, v in stats["by_mod"].items():
        lines.append(f"    {k}: {v}")
    lines.append("  by race:  " + ", ".join(f"{k}={v}" for k, v in stats["by_race"].items()))
    lines.append("  by gender:" + ", ".join(f"{k}={v}" for k, v in stats["by_gender"].items()))
    lines.append("")

    lines.append(f"RENAMES ({len(renames)})")
    for old, new in renames:
        lines.append(f"  {old}  ->  {new.name}")
    lines.append("")

    if collisions:
        lines.append(f"!! NAME COLLISIONS ({len(collisions)}) -- NOT renamed, fix by hand")
        for old, new in collisions:
            lines.append(f"  {old}  would collide with existing  {new}")
        lines.append("")

    mismatch_lines = [w for w in warnings if w.startswith("MISMATCH")]
    other_warn = [w for w in warnings if not w.startswith("MISMATCH")]
    lines.append(f"MISMATCHED FILES ({len(mismatch_lines)}) -- fix folder or filename so they agree")
    for w in mismatch_lines:
        lines.append("  " + w)
    lines.append("")

    if other_warn:
        lines.append(f"OTHER WARNINGS ({len(other_warn)})")
        for w in other_warn:
            lines.append("  " + w)
        lines.append("")

    if skipped:
        lines.append(f"SKIPPED ({len(skipped)})")
        for kind, rel in skipped:
            lines.append(f"  [{kind}] {rel}")
        lines.append("")

    path.write_text("\n".join(lines), encoding="utf-8")


def run_sanitize(root, manifest, report, url_base=None, apply=True):
    """Rename unsafe image/folder names and write previews.json."""
    args = SimpleNamespace(root=str(root), manifest=str(manifest),
                           report=str(report), url_base=url_base, apply=apply)
    root = Path(args.root).resolve()
    if not root.is_dir():
        print(f"ERROR: --root '{root}' is not a directory", file=sys.stderr)
        return 2

    renames, collisions = plan_renames(root)

    if args.apply:
        if collisions:
            print(f"ERROR: {len(collisions)} name collision(s) would occur; resolve them first (see report). Nothing renamed.", file=sys.stderr)
        else:
            n = apply_renames(renames)
            print(f"Renamed {n} item(s).")

    records, warnings, skipped, stats = build_manifest(root, args.url_base)

    manifest = {
        "meta": {
            "generated": datetime.now(timezone.utc).isoformat(timespec="seconds"),
            "root_folder_note": "paths are relative to the images root and exclude its folder name; compose URLs as <base>/<your images folder>/<file>",
            "url_base": args.url_base,
            "counts": stats,
        },
        "images": records,
    }

    if args.apply and not collisions:
        Path(args.manifest).write_text(json.dumps(manifest, indent=2, ensure_ascii=True), encoding="utf-8")
        write_report(Path(args.report), renames, collisions, records, warnings, skipped, stats, applied=True)
        print(f"Wrote {args.manifest} ({stats['total_images']} images) and {args.report}.")
    else:
        # Dry run: write only the human report so nothing is committed by accident.
        write_report(Path(args.report), renames, collisions, records, warnings, skipped, stats, applied=False)
        print("DRY RUN -- no files renamed, no manifest written.")
        print(f"  would rename : {len(renames)}")
        print(f"  collisions   : {len(collisions)}")
        print(f"  images       : {stats['total_images']}")
        print(f"  mismatches   : {stats['mismatches']}")
        print(f"Preview details written to {args.report}. Re-run with --apply to commit.")

    return 0


# ===========================================================================
# PART 2 -- spreadsheet export
# (was export_data.py)
# ===========================================================================
KNOWN_RACE_TYPES = {"Argonian", "Elf", "Human", "HumanChild", "Khajiit", "Orc",
                    "COTR", "UBE"}
BRACKETS = "()[]{}"

# --- Display config (edit here, not in the HTML) ---------------------------
# Friendly display names for individual hairstyles (does not change previewKey/spell).
NAME_OVERRIDES = {
    "Valkyr - CDante": "Valkyr - Classic Dante",
    "Valkyr - Dany":   "Valkyr - Daenerys",
}
# The dropdown/browser group label comes straight from the sheet's Mod column.
# Add an entry here only to show a different label than the Mod value.
MOD_TO_PACK = {}
def pack_for(mod, preview_key):
    return MOD_TO_PACK.get(mod, mod)

# Requirements now live on the "Hairstyle Requirements" tab (ID / Link Text /
# Link), referenced by ID from the Hairstyles sheet. Tab row order is display
# order. Nothing about them is hardcoded here any more.
REQ_SHEET = "Hairstyle Requirements"
# --------------------------------------------------------------------------


def preview_key(name: str) -> str:
    """Spreadsheet Hairstyle Name -> filename hairstyle key.

    Must stay identical to the Sheets formula that names the screenshots:
    drop spaces, apostrophes and brackets, then turn EVERY hyphen into an
    underscore. That last step matters for names with an internal hyphen
    ("Ronin Top-knot" -> Vanilla_RoninTop_knot), not just the " - " separator.
    """
    s = name.replace(" - ", "_")
    for ch in [" ", "'"] + list(BRACKETS):
        s = s.replace(ch, "")
    return s.replace("-", "_")


def split_categories(cell):
    """Parse the 'Race_Gender Categories' cell into ({race: [genders]}, bad[]).

    Entries look like "Human_Female" or "HumanChild_Male" -- split on the LAST
    underscore, since race types may contain one (HumanChild) but genders never
    do. This column is normally an array formula producing the cross product of
    Valid Races x Valid Genders; replace it with a literal list on rows where
    the hairstyle isn't symmetric (e.g. male elves + female humans only).
    """
    vmap, bad = {}, []
    if cell is None:
        return vmap, bad
    for part in str(cell).split(","):
        p = part.strip()
        if not p or p.lower() == "none":
            continue
        if "_" not in p:
            bad.append(p)
            continue
        race, gender = p.rsplit("_", 1)
        race, gender = race.strip(), gender.strip()
        if not race or not gender:
            bad.append(p)
            continue
        vmap.setdefault(race, [])
        if gender not in vmap[race]:
            vmap[race].append(gender)
    return vmap, bad


def split_semi(cell) -> list:
    """Semicolon- (or comma-) separated EditorID list from one cell."""
    if not cell:
        return []
    out = []
    for part in str(cell).replace(",", ";").split(";"):
        p = part.strip()
        if p and p not in out:
            out.append(p)
    return out


def derive_descendants(npcs):
    """Work out, for every template, which descendants must be excluded from its
    wig and which share its face.

    A descendant keeps the root's face only if EVERY link in the chain has Use
    Traits. The first record on each branch where that breaks is the one to
    exclude -- excluding it also excludes everything below it, so deeper nodes
    don't need listing.
    """
    by_eid = {n["editorid"]: n for n in npcs}
    kids = defaultdict(list)
    for n in npcs:
        if n.get("template"):
            kids[n["template"]].append(n["editorid"])

    for n in npcs:
        exclude, sharing = [], []
        stack = [(k, True) for k in kids.get(n["editorid"], [])]
        seen = set()
        while stack:
            eid, chain_intact = stack.pop()
            if eid in seen:            # guard against a cyclic template chain
                continue
            seen.add(eid)
            child = by_eid.get(eid)
            if chain_intact and child and child.get("templateUseTraits"):
                sharing.append(eid)
                stack.extend((gk, True) for gk in kids.get(eid, []))
            else:
                exclude.append(eid)    # its own descendants follow it
        if exclude:
            n["excludeDescendants"] = sorted(exclude)
        if sharing:
            n["faceSharingDescendants"] = sorted(sharing)
    return by_eid


def build_valid_map(races: list, genders: list, categories: dict) -> dict:
    """Per-race gender validity.

    The categories column wins when present (it can express asymmetric combos);
    otherwise fall back to the plain cross product of Valid Races x Valid
    Genders so rows without that column still work.
    """
    if categories:
        return {r: list(g) for r, g in categories.items() if g}
    return {r: list(genders) for r in races if genders}


def split_ids(cell) -> list:
    """Parse the Requirements cell: comma-separated requirement IDs."""
    if cell is None:
        return []
    out = []
    for part in str(cell).split(","):
        p = part.strip()
        if p and p.lower() != "none" and p not in out:
            out.append(p)
    return out


def split_cond_reqs(cell):
    """Parse the 'Race/Gender Specific Requirements' cell.

    Entries are "Condition:RequirementID", where Condition is either a race
    type ("Khajiit:KS_SMP_Khajiit") or a Race_Gender combo
    ("Khajiit_Female:Some_ID"). Commas separate entries. Returns
    ({condition: [ids]}, bad[]).
    """
    out, bad = {}, []
    if cell is None:
        return out, bad
    for part in str(cell).split(","):
        p = part.strip()
        if not p or p.lower() == "none":
            continue
        if ":" not in p:
            bad.append(p)
            continue
        cond, rid = p.split(":", 1)
        cond, rid = cond.strip(), rid.strip()
        if not cond or not rid:
            bad.append(p)
            continue
        out.setdefault(cond, [])
        if rid not in out[cond]:
            out[cond].append(rid)
    return out, bad


def cond_race(cond: str) -> str:
    """The race a condition applies to ("Khajiit" or "Khajiit_Female" -> Khajiit)."""
    if cond in KNOWN_RACE_TYPES:
        return cond
    if "_" in cond:
        base, last = cond.rsplit("_", 1)
        if last in KNOWN_GENDERS:
            return base
    return cond


def read_requirements(wb):
    """Read the requirements tab into an ordered list of {id, text, url}."""
    if REQ_SHEET not in wb.sheetnames:
        return [], [f"missing sheet {REQ_SHEET!r}"]
    ws = wb[REQ_SHEET]
    reqs, problems, seen = [], [], set()
    for r in range(2, ws.max_row + 1):
        rid = ws.cell(r, 1).value
        if rid is None or not str(rid).strip():
            continue
        rid = str(rid).strip()
        text = ws.cell(r, 2).value
        url = ws.cell(r, 3).value
        text = str(text).strip() if text else ""
        url = str(url).strip() if url else ""
        if rid in seen:
            problems.append(f"duplicate requirement ID {rid!r} (row {r})")
            continue
        seen.add(rid)
        if not text:
            problems.append(f"requirement {rid!r} has no link text")
        if not url:
            problems.append(f"requirement {rid!r} has no link")
        reqs.append({"id": rid, "text": text or rid, "url": url})
    return reqs, problems


def split_list(cell) -> list:
    """Split a comma-separated cell into a clean list, dropping blanks/None."""
    if cell is None:
        return []
    out = []
    for part in str(cell).split(","):
        p = part.strip()
        if p and p.lower() != "none":
            out.append(p)
    return out


def load_sheet(wb, name):
    if name not in wb.sheetnames:
        print(f"ERROR: sheet '{name}' not found. Sheets present: {wb.sheetnames}", file=sys.stderr)
        raise SystemExit(2)
    return wb[name]


def run_export(xlsx, previews, outdir, report):
    """Rebuild Data/*.json from the spreadsheet and the manifest."""
    args = SimpleNamespace(xlsx=str(xlsx), previews=str(previews),
                           outdir=str(outdir), report=str(report))
    outdir = Path(args.outdir)
    outdir.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.load_workbook(args.xlsx, data_only=True)

    # ---- Race map (first two columns) ----
    # ---- Races (Race Group | Race EditorID | Race Name | Race Type) ----
    # Sheet order is preserved: it drives both the group headings and the race
    # order in the tool's race dropdown.
    ws = load_sheet(wb, "Races")
    race_list = []          # ordered, as written to races.json
    race_map = {}           # EditorID -> race type, for resolving NPCs
    dupe_races = []
    for r in range(2, ws.max_row + 1):
        group = ws.cell(r, 1).value
        eid = ws.cell(r, 2).value
        name = ws.cell(r, 3).value
        rtype = ws.cell(r, 4).value
        if not eid:
            continue
        eid = str(eid).strip()
        if eid in race_map:
            dupe_races.append(eid)
            continue
        rtype = str(rtype).strip() if rtype else None
        race_map[eid] = rtype
        race_list.append({
            "editorid": eid,
            "name": str(name).strip() if name else eid,
            "group": str(group).strip() if group else "",
            "type": rtype,
        })

    # ---- Hairstyles ----
    ws = load_sheet(wb, "Hairstyles")
    hairstyles = []
    key_to_rows = defaultdict(list)   # detect collisions
    empty_valid = []
    bad_pairs = []                    # malformed entries in the exceptions column
    bad_cond_entries = []             # malformed entries in the conditional-req column
    asym = []                         # rows using per-race exceptions (for the report)
    hair_editorid_to_style = {}       # in-game hair EditorID -> hairstyle rec
    bald_style = None
    for r in range(2, ws.max_row + 1):
        mod = ws.cell(r, 1).value
        editor_ids = ws.cell(r, 2).value
        name = ws.cell(r, 3).value
        spell = ws.cell(r, 6).value
        if not (name and spell):
            continue
        name = str(name).strip()
        races = split_list(ws.cell(r, 7).value)
        genders = split_list(ws.cell(r, 8).value)
        categories, bad_cats = split_categories(ws.cell(r, 9).value)   # col 9
        for bad in bad_cats:
            bad_pairs.append((name, bad))
        reqs_ids = split_ids(ws.cell(r, 10).value)                     # col 10
        cond_reqs, bad_conds = split_cond_reqs(ws.cell(r, 11).value)   # col 11
        for bad in bad_conds:
            bad_cond_entries.append((name, bad))
        valid_map = build_valid_map(races, genders, categories)
        # Flat lists stay in sync with the map so older consumers still work.
        races = sorted(valid_map.keys())
        genders = sorted({g for gs in valid_map.values() for g in gs})
        key = preview_key(name)
        rec = {
            "mod": str(mod).strip() if mod else "",
            "name": name,
            "spell": str(spell).strip(),
            "previewKey": key,
            "validRaceTypes": races,
            "validGenders": genders,
            "validGendersByRace": valid_map,
            "requirements": reqs_ids,
            "condRequirements": cond_reqs,
            "editorIDs": split_list(editor_ids),
            "pack": pack_for(str(mod).strip() if mod else "", key),
            "displayName": NAME_OVERRIDES.get(name, name),
        }
        hairstyles.append(rec)
        key_to_rows[key].append(name)
        # Flag rows whose combos aren't the plain cross product -- these are the
        # asymmetric ones only the categories column can express.
        if valid_map and set(valid_map) and any(
                sorted(gs) != sorted(genders) for gs in valid_map.values()):
            asym.append((name, valid_map))
        if not races or not genders:
            empty_valid.append((name, races, genders))
        # index each in-game hair EditorID so NPC vanilla hair can be resolved
        for hid in split_list(editor_ids):
            hair_editorid_to_style.setdefault(hid, rec)
        if name.lower().endswith("bald") or key.lower().endswith("bald"):
            bald_style = rec

    def resolve_default(vanilla_hair):
        """Map an NPC's Vanilla Hair ID to a hairstyle rec (or None)."""
        if not vanilla_hair:
            return None
        vh = str(vanilla_hair).strip()
        if vh in hair_editorid_to_style:
            return hair_editorid_to_style[vh]
        if vh.lower() in ("(none/bald)", "none/bald", "bald", "none"):
            return bald_style
        return None

    # ---- NPCs ----
    ws = load_sheet(wb, "NPC List")
    # Read by header name, not position: the NPC List column order now mirrors
    # the xEdit CSV export and may be rearranged again.
    hdr = {}
    for c in range(1, ws.max_column + 1):
        h = ws.cell(1, c).value
        if h:
            hdr[str(h).strip().lower()] = c
    def cell(r, *names, default=""):
        for n in names:
            c = hdr.get(n)
            if c:
                v = ws.cell(r, c).value
                if v is not None and str(v).strip() != "":
                    return str(v).strip()
        return default
    def flag(r, *names):
        """A template flag column: 1/true/yes means set."""
        return cell(r, *names).lower() in ("1", "true", "yes", "y", "x")

    missing_cols = [n for n in ("editorid", "race", "gender") if n not in hdr]
    if missing_cols:
        print(f"ERROR: NPC List is missing column(s): {missing_cols}", file=sys.stderr)
        return 2

    npcs = []
    npc_race_missing = set()
    for r in range(2, ws.max_row + 1):
        eid = cell(r, "editorid", "npc editorid")
        if not eid:
            continue
        race = cell(r, "race", "npc race")
        rtype = race_map.get(race)
        if race and race not in race_map:
            npc_race_missing.add(race)
        vanilla_hair = cell(r, "hairstyle", "vanilla hair id", "vanillahair")
        default_style = resolve_default(vanilla_hair)
        rec = {
            "formid": cell(r, "formid"),
            "editorid": eid,
            "name": cell(r, "name", "npc name"),
            "race": race,
            "raceType": rtype,
            "gender": cell(r, "gender", "npc gender"),
            "vanillaHair": vanilla_hair,
            "defaultPreviewKey": default_style["previewKey"] if default_style else None,
            "defaultHairstyleName": default_style["name"] if default_style else None,
            "defaultSpell": default_style["spell"] if default_style else None,
        }
        # Template data. SPID hands a template's spells to every descendant, so a
        # wig on a template needs explicit exclusions for the ones that don't
        # share its face. Use Traits is what decides that.
        template = cell(r, "template")
        if template:
            rec["template"] = template
            if flag(r, "templateusetraits"): rec["templateUseTraits"] = 1
        # The sheet's own lists are kept only to check the derivation below.
        rec["_sheetExclude"] = split_semi(cell(r, "excludedescendants"))
        rec["_sheetSharing"] = split_semi(cell(r, "facesharingdescendants"))
        npcs.append(rec)

    # Guard against the same record appearing twice with different template data.
    seen_npc, dupes, dupe_conflicts = {}, [], []
    unique = []
    for n in npcs:
        prev = seen_npc.get(n["editorid"])
        if prev is None:
            seen_npc[n["editorid"]] = n
            unique.append(n)
            continue
        dupes.append(n["editorid"])
        if any(prev.get(k) != n.get(k) for k in
               ("template", "templateUseTraits", "_sheetExclude", "_sheetSharing",
                "race", "gender", "vanillaHair")):
            dupe_conflicts.append(n["editorid"])
    npcs = unique
    derive_descendants(npcs)

    # ---- Previews: build key -> {race: {gender: file}} ----
    previews = json.loads(Path(args.previews).read_text(encoding="utf-8"))
    pv_index = defaultdict(lambda: defaultdict(dict))
    pv_keys = set()
    for img in previews["images"]:
        pv_index[img["hairstyle"]][img["race"]][img["gender"]] = img["file"]
        pv_keys.add(img["hairstyle"])

    # ---- Validation ----
    hs_keys = {h["previewKey"] for h in hairstyles}
    no_preview = sorted(k for k in hs_keys if k not in pv_keys)
    orphan_imgs = sorted(pv_keys - hs_keys)
    collisions = {k: v for k, v in key_to_rows.items() if len(v) > 1}
    bad_race_types = sorted({t for h in hairstyles for t in h["validRaceTypes"]
                             if t not in KNOWN_RACE_TYPES})
    bad_genders = sorted({g for h in hairstyles for g in h["validGenders"]
                          if g not in KNOWN_GENDERS})

    # ---- Write JSON ----
    (outdir / "races.json").write_text(json.dumps(race_list, indent=2, ensure_ascii=True), encoding="utf-8")
    (outdir / "hairstyles.json").write_text(json.dumps(hairstyles, indent=2, ensure_ascii=True), encoding="utf-8")
    # Keep the sheet's own lists for the check further down, then drop them so
    # they never reach the output.
    sheet_lists = {n["editorid"]: (n.pop("_sheetExclude", []), n.pop("_sheetSharing", []))
                   for n in npcs}
    (outdir / "npcs.json").write_text(json.dumps(npcs, indent=2, ensure_ascii=True), encoding="utf-8")
    requirements, req_problems = read_requirements(wb)
    req_ids = {r["id"] for r in requirements}
    used_ids = Counter()
    undefined_refs = []      # (hairstyle, id)
    dead_conditions = []     # (hairstyle, condition, id, valid races)
    for h in hairstyles:
        for rid in h["requirements"]:
            used_ids[rid] += 1
            if rid not in req_ids:
                undefined_refs.append((h["name"], rid))
        valid_races = set(h["validGendersByRace"].keys())
        for cond, ids in h["condRequirements"].items():
            for rid in ids:
                used_ids[rid] += 1
                if rid not in req_ids:
                    undefined_refs.append((h["name"], rid))
                if cond_race(cond) not in valid_races:
                    dead_conditions.append((h["name"], cond, rid, sorted(valid_races)))
    unused_ids = sorted(req_ids - set(used_ids))
    no_reqs = [h["name"] for h in hairstyles if not h["requirements"]]
    (outdir / "requirements.json").write_text(
        json.dumps(requirements, indent=2, ensure_ascii=True), encoding="utf-8")

    # ---- Report ----
    lines = []
    lines.append("Export report")
    lines.append("=============")
    lines.append(f"NPCs exported          : {len(npcs)}")
    lines.append(f"Hairstyles exported    : {len(hairstyles)}")
    lines.append(f"Races in map           : {len(race_map)}")
    n_tmpl = sum(1 for n in npcs if n.get("excludeDescendants") or n.get("faceSharingDescendants"))
    n_child = sum(1 for n in npcs if n.get("template"))
    n_excl = sum(len(n.get("excludeDescendants", [])) for n in npcs)
    n_share = sum(len(n.get("faceSharingDescendants", [])) for n in npcs)
    biggest = max((len(n.get("excludeDescendants", [])) for n in npcs), default=0)
    lines.append(f"Template records       : {n_tmpl}")
    lines.append(f"NPCs using a template  : {n_child}")
    lines.append(f"Duplicate NPC rows dropped: {len(dupes)}"
                 + (f"  ({len(dupe_conflicts)} disagreed with the first copy: "
                    f"{sorted(set(dupe_conflicts))[:5]})" if dupe_conflicts else ""))
    lines.append(f"Exclusions to write    : {n_excl}  (largest single list: {biggest})")
    lines.append(f"Face-sharing links     : {n_share}")
    lines.append(f"Preview images         : {len(previews['images'])}  ({len(pv_keys)} unique keys)")
    lines.append("")
    lines.append(f"Hairstyles with NO preview image ({len(no_preview)}) -> tool shows a placeholder:")
    for k in no_preview:
        lines.append(f"   {k}   (from: {', '.join(key_to_rows[k])})")
    lines.append("")
    lines.append(f"Preview images with NO hairstyle row ({len(orphan_imgs)}):")
    for k in orphan_imgs:
        lines.append(f"   {k}")
    lines.append("")
    lines.append(f"Hairstyle rows with empty/None valid race or gender ({len(empty_valid)}):")
    for name, rc, ge in empty_valid:
        lines.append(f"   {name}   races={rc}  genders={ge}")
    lines.append("")
    lines.append(f"Preview-key collisions ({len(collisions)}):")
    for k, v in collisions.items():
        lines.append(f"   {k}: {v}")
    lines.append("")
    lines.append(f"Hairstyles with asymmetric race/gender combos ({len(asym)}):")
    for name, vmap in asym:
        lines.append(f"   {name}: " + "; ".join(f"{r}={'/'.join(g)}" for r, g in sorted(vmap.items())))
    lines.append("")
    lines.append(f"Malformed entries in the Race_Gender Categories column ({len(bad_pairs)}) -- expected Race_Gender:")
    for name, bad in bad_pairs:
        lines.append(f"   {name}: {bad!r}")
    lines.append("")
    lines.append(f"Requirements defined on the {REQ_SHEET!r} tab: {len(requirements)}")
    for p in req_problems:
        lines.append(f"   PROBLEM: {p}")
    lines.append("")
    lines.append(f"Requirement IDs referenced but NOT defined ({len(undefined_refs)}):")
    for name, rid in undefined_refs:
        lines.append(f"   {name}: {rid!r}")
    lines.append("")
    lines.append(f"Requirements defined but never used ({len(unused_ids)}): {unused_ids}")
    lines.append("")
    lines.append(f"Hairstyles with NO requirement at all ({len(no_reqs)}):")
    for n in no_reqs:
        lines.append(f"   {n}")
    lines.append("")
    lines.append(f"Conditional requirements for a race the hairstyle isn't valid for ({len(dead_conditions)}):")
    for name, cond, rid, vr in dead_conditions:
        lines.append(f"   {name}: {cond}:{rid}  (valid races: {vr})")
    lines.append("")
    lines.append(f"Malformed entries in the Race/Gender Specific Requirements column ({len(bad_cond_entries)}) -- expected Condition:RequirementID:")
    for name, bad in bad_cond_entries:
        lines.append(f"   {name}: {bad!r}")
    lines.append("")
    groups_seen = []
    for rec in race_list:
        if rec["group"] not in groups_seen:
            groups_seen.append(rec["group"])
    # The spreadsheet carries its own descendant lists; the tool uses the derived
    # ones, but disagreement means the xEdit script and this script see the tree
    # differently -- worth knowing, because CSV imports rely on those columns.
    miss_ex, extra_ex, miss_sh, extra_sh, bad_rows = 0, 0, 0, 0, []
    for n in npcs:
        sh_ex, sh_sh = sheet_lists.get(n["editorid"], ([], []))
        d_ex, s_ex = set(n.get("excludeDescendants", [])), set(sh_ex)
        d_sh, s_sh = set(n.get("faceSharingDescendants", [])), set(sh_sh)
        if d_ex != s_ex or d_sh != s_sh:
            bad_rows.append((n["editorid"], sorted(d_ex - s_ex), sorted(s_ex - d_ex),
                             sorted(d_sh - s_sh), sorted(s_sh - d_sh)))
        miss_ex += len(d_ex - s_ex); extra_ex += len(s_ex - d_ex)
        miss_sh += len(d_sh - s_sh); extra_sh += len(s_sh - d_sh)
    lines.append(f"Rows where the sheet's descendant lists disagree with the derived ones "
                 f"({len(bad_rows)}):")
    lines.append(f"   exclusions   : {miss_ex} missing from the sheet, {extra_ex} extra")
    lines.append(f"   face-sharing : {miss_sh} missing from the sheet, {extra_sh} extra")
    for eid, mex, xex, msh, xsh in bad_rows[:25]:
        bits = []
        if mex: bits.append(f"exclude missing {mex[:5]}")
        if xex: bits.append(f"exclude extra {xex[:5]}")
        if msh: bits.append(f"sharing missing {msh[:5]}")
        if xsh: bits.append(f"sharing extra {xsh[:5]}")
        lines.append(f"   {eid}: " + "; ".join(bits))
    if len(bad_rows) > 25:
        lines.append(f"   ... and {len(bad_rows) - 25} more")
    lines.append("")
    lines.append(f"Race groups (dropdown order): {groups_seen}")
    lines.append(f"Duplicate race EditorIDs ({len(dupe_races)}): {sorted(set(dupe_races))}")
    no_name = [r["editorid"] for r in race_list if r["name"] == r["editorid"]]
    lines.append(f"Races with no display name ({len(no_name)}): {no_name}")
    bad_map_types = sorted({t for t in race_map.values()
                            if t and t not in KNOWN_RACE_TYPES})
    lines.append(f"Unknown race types on the race map ({len(bad_map_types)}): {bad_map_types}")
    lines.append(f"Unknown gender values ({len(bad_genders)}): {bad_genders}")
    lines.append(f"Unknown valid-race-type values ({len(bad_race_types)}): {bad_race_types}")
    lines.append(f"NPC races missing from race map ({len(npc_race_missing)}): {sorted(npc_race_missing)}")
    report = "\n".join(lines)
    Path(args.report).write_text(report, encoding="utf-8")
    print(report)
    print(f"\nWrote {outdir/'npcs.json'}, {outdir/'hairstyles.json'}, {outdir/'races.json'} and {args.report}.")
    return 0


# ===========================================================================
# PART 3 -- refresh the sample embedded in index.html
# ===========================================================================

def rebuild_sample(data_dir, manifest_path, index_path):
    """Replace the EMBEDDED_DATA blob in index.html with fresh sample data."""
    if not index_path.exists():
        print(f"  skipped: {index_path.name} not found")
        return
    html = index_path.read_text(encoding="utf-8")
    pattern = re.compile(r"(<script>const EMBEDDED_DATA = )(.*?)(;</script>)", re.S)
    if not pattern.search(html):
        print("  skipped: no EMBEDDED_DATA block in index.html")
        return

    npcs = json.loads((data_dir / "npcs.json").read_text(encoding="utf-8"))
    hairs = json.loads((data_dir / "hairstyles.json").read_text(encoding="utf-8"))
    races = json.loads((data_dir / "races.json").read_text(encoding="utf-8"))
    reqs = json.loads((data_dir / "requirements.json").read_text(encoding="utf-8"))
    previews = json.loads(manifest_path.read_text(encoding="utf-8"))
    previews = previews.get("images", previews)

    by_group = defaultdict(list)
    for n in npcs:
        by_group[(n.get("raceType"), n.get("gender"))].append(n)
    sample_npcs = []
    for key in sorted(by_group, key=lambda k: (str(k[0]), str(k[1]))):
        sample_npcs.extend(by_group[key][:SAMPLE_NPCS_PER_GROUP])

    by_pv = defaultdict(list)
    for p in previews:
        by_pv[(p["mod"], p["race"], p["gender"])].append(p)
    sample_pv = []
    for key in sorted(by_pv):
        sample_pv.extend(by_pv[key][:SAMPLE_PREVIEWS_PER_MOD])

    blob = {
        "npcs": sample_npcs,
        "hairstyles": hairs,      # all of them; the browser needs the full list
        "races": races,
        "previews": sample_pv,
        "requirements": reqs,
    }
    payload = json.dumps(blob, ensure_ascii=True, separators=(",", ":"))
    index_path.write_text(pattern.sub(lambda m: m.group(1) + payload + m.group(3),
                                      html, count=1), encoding="utf-8")
    print(f"  embedded sample: {len(sample_npcs)} NPCs, {len(hairs)} hairstyles, "
          f"{len(sample_pv)} previews, {len(reqs)} requirements")


# ===========================================================================
# Orchestration
# ===========================================================================

# ===========================================================================
# Report handling: keep logs out of the repo unless they're asked for
# ===========================================================================

# Lines that mean something is actually wrong and worth reading the log over.
PROBLEM_MARKERS = [
    "collision", "Malformed", "NOT defined", "no display name",
    "Duplicate race", "isn't valid for", "empty/None valid race",
    "NO requirement at all", "missing from race map", "Unknown gender",
    "Unknown valid-race-type", "Unknown race types", "mismatch", "PROBLEM",
]
# Lines that are just worth knowing about (e.g. screenshots not taken yet).
NOTE_MARKERS = ["NO preview image", "NO hairstyle row"]


def scan_report(text):
    """Return (problems, notes) as lists of 'headline (count)' strings."""
    problems, notes = [], []
    for line in (text or "").splitlines():
        stripped = line.strip()
        if not stripped or stripped.startswith(("   ", "\t")) or line.startswith("   "):
            continue
        count = None
        m = re.search(r"\((\d+)\)", stripped)
        if m:
            count = int(m.group(1))
        else:
            m2 = re.search(r":\s*\[(.*)\]\s*$", stripped)      # e.g. "Unknown gender values (0): []"
            if m2:
                count = 0 if not m2.group(1).strip() else 1
        if count is None or count == 0:
            continue
        headline = stripped.split(" -- ")[0].rstrip(":")
        if any(k.lower() in stripped.lower() for k in PROBLEM_MARKERS):
            problems.append(headline)
        elif any(k.lower() in stripped.lower() for k in NOTE_MARKERS):
            notes.append(headline)
    return problems, notes


def ask_save_reports(problems):
    """Enter saves nothing; y then Enter writes the logs."""
    if not sys.stdin or not sys.stdin.isatty():
        return False                     # non-interactive run: never leave files behind
    hint = "RECOMMENDED - something needs attention" if problems else "not needed"
    try:
        answer = input(f"\nWrite reports to disk? [y/N] ({hint}): ")
    except (EOFError, KeyboardInterrupt):
        print()
        return False
    return answer.strip().lower().startswith("y")


def main():
    xlsx = rel(XLSX)
    images = rel(IMAGES_ROOT)
    manifest = rel(MANIFEST)
    data_dir = rel(DATA_DIR)
    index_html = rel(INDEX_HTML)
    # Reports are written here first; they only reach the repo folder if asked for.
    tmpdir = Path(tempfile.mkdtemp(prefix="hdf_reports_"))
    preview_report = tmpdir / PREVIEW_REPORT
    export_report = tmpdir / EXPORT_REPORT

    if not xlsx.exists():
        sys.exit(f"ERROR: spreadsheet not found: {xlsx}")
    if not images.is_dir():
        sys.exit(f"ERROR: images folder not found: {images}")

    print("=" * 70)
    print("STEP 1/3  Sanitizing image names and rebuilding the manifest")
    print("=" * 70)
    before = manifest.stat().st_mtime if manifest.exists() else None
    code = run_sanitize(images, manifest, preview_report, IMAGE_URL_BASE)
    if code not in (0, None):
        sys.exit(f"\nSanitize step failed (exit {code}).")

    # The sanitizer refuses to write the manifest if two files would collide.
    # Catch that here so the export can't silently run on a stale manifest.
    if not manifest.exists():
        sys.exit(f"\nERROR: {manifest.name} was not written. See "
                 f"{preview_report.name} and fix the problem, then re-run.")
    if before is not None and manifest.stat().st_mtime == before:
        sys.exit(f"\nERROR: {manifest.name} was NOT updated (likely a name collision). "
                 f"See {preview_report.name}. Stopping so the export can't use stale data.")

    print()
    print("=" * 70)
    print("STEP 2/3  Exporting spreadsheet data")
    print("=" * 70)
    code = run_export(xlsx, manifest, data_dir, export_report)
    if code not in (0, None):
        sys.exit(f"\nExport step failed (exit {code}).")

    print()
    print("=" * 70)
    print("STEP 3/3  Refreshing the sample embedded in index.html")
    print("=" * 70)
    if REBUILD_EMBEDDED_SAMPLE:
        rebuild_sample(data_dir, manifest, index_html)
    else:
        print("  disabled (REBUILD_EMBEDDED_SAMPLE = False)")

    print()
    print("=" * 70)
    print("Done.")
    print(f"  manifest : {manifest}")
    print(f"  data     : {data_dir}")

    # ---- what, if anything, needs attention -------------------------------
    texts = []
    for rp in (preview_report, export_report):
        if rp.exists():
            texts.append(rp.read_text(encoding="utf-8", errors="replace"))
    problems, notes = scan_report("\n".join(texts))
    if problems:
        print()
        print("  SOMETHING NEEDS ATTENTION:")
        for line in problems:
            print(f"    - {line}")
    if notes:
        print()
        print("  Worth knowing:")
        for line in notes:
            print(f"    - {line}")
    if not problems and not notes:
        print("  No problems found.")

    # ---- offer the logs ---------------------------------------------------
    saved = []
    if ask_save_reports(problems):
        for rp in (preview_report, export_report):
            if rp.exists():
                dest = rel(rp.name)
                dest.write_text(rp.read_text(encoding="utf-8", errors="replace"),
                                encoding="utf-8")
                saved.append(dest)
    shutil.rmtree(tmpdir, ignore_errors=True)
    if saved:
        print("\nReports written:")
        for d in saved:
            print(f"  {d}")
        print("(These are ignored by git if you add *_report.txt to .gitignore.)")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
