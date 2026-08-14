#!/usr/bin/env python3
"""
export_data.py
==============

Turns the maintainer spreadsheet into the JSON the wig-assignment tool reads,
and validates it against previews.json. Re-run this any time you edit the
spreadsheet -- the spreadsheet stays the single source of truth.

Reads three sheets:
  * "NPC List"               -> npcs.json
  * "Hairstyles"             -> hairstyles.json
  * "Race Default Hairstyles"-> races.json   (first two columns only)

The preview key for each hairstyle is derived from its Hairstyle Name using the
transform we validated: replace " - " with "_", then delete spaces, apostrophes
and brackets. That key is how the tool finds the matching screenshot in
previews.json.

Hairstyles sheet columns read:
  1 Mod | 2 in-game hair EditorIDs | 3 Name | 6 Spell
  7 Valid Races | 8 Valid Genders | 9 Race/Gender Exceptions (OPTIONAL)

Column 9 ("Race_Gender Categories") is the authoritative validity list. It is
normally an array formula producing the cross product of Valid Races x Valid
Genders, so symmetric rows need no attention. To express a hairstyle that ISN'T
symmetric, replace that cell's formula with a literal list:

  "Elf_Male, Human_Male, Human_Female, Orc_Male, Orc_Female, Khajiit_Female"
  -> male elves, all humans and orcs, and female khajiit only.

Entries split on the LAST underscore, so HumanChild_Male parses correctly. If
column 9 is empty the exporter falls back to the cross product of columns 7-8.
Columns 7 and 8 are still emitted (derived from the final combo list) and are
what the "Valid Races"/"Valid Genders" filters use.

Outputs (JSON, ASCII-safe):
  npcs.json       [{formid, editorid, name, race, raceType, gender, vanillaHair}]
  hairstyles.json [{mod, name, spell, previewKey, validRaceTypes[], validGenders[],
                    validGendersByRace{}}]
  races.json      {race: raceType}

Validation (printed, and written to export_report.txt):
  - hairstyles whose preview key has no image (shown as placeholders in the tool)
  - preview keys with no hairstyle row (orphan screenshots)
  - hairstyle rows with empty/None valid race types or genders
  - NPC races missing from the race map
  - preview-key collisions

Usage:
  python export_data.py --xlsx Skyrim_NPC_Hairstyle_Distribution.xlsx --previews previews.json --outdir data
Requires: openpyxl  (pip install openpyxl)
"""

import argparse
import json
import os
import sys
from pathlib import Path
from collections import defaultdict, Counter

try:
    import openpyxl
except ImportError:
    print("This script needs openpyxl.  Install it with:  pip install openpyxl", file=sys.stderr)
    raise SystemExit(1)

KNOWN_RACE_TYPES = {"Argonian", "Elf", "Human", "HumanChild", "Khajiit", "Orc"}
KNOWN_GENDERS = {"Male", "Female"}
BRACKETS = "()[]{}"

# --- Display config (edit here, not in the HTML) ---------------------------
# Friendly display names for individual hairstyles (does not change previewKey/spell).
NAME_OVERRIDES = {
    "Valkyr - CDante": "Valkyr - Classic Dante",
    "Valkyr - Dany":   "Valkyr - Daenerys",
}
# The dropdown/browser group label comes straight from the sheet's Mod column.
# Add an entry here only to show a different label than the Mod value.
MOD_TO_PACK = {}
def pack_for(mod, preview_key):
    return MOD_TO_PACK.get(mod, mod)

# Requirements now live on the "Hairstyle Requirements" tab (ID / Link Text /
# Link), referenced by ID from the Hairstyles sheet. Tab row order is display
# order. Nothing about them is hardcoded here any more.
REQ_SHEET = "Hairstyle Requirements"
# --------------------------------------------------------------------------


def preview_key(name: str) -> str:
    """Spreadsheet Hairstyle Name -> filename hairstyle key. Validated transform."""
    s = name.replace(" - ", "_")
    for ch in [" ", "'"] + list(BRACKETS):
        s = s.replace(ch, "")
    return s


def split_categories(cell):
    """Parse the 'Race_Gender Categories' cell into ({race: [genders]}, bad[]).

    Entries look like "Human_Female" or "HumanChild_Male" -- split on the LAST
    underscore, since race types may contain one (HumanChild) but genders never
    do. This column is normally an array formula producing the cross product of
    Valid Races x Valid Genders; replace it with a literal list on rows where
    the hairstyle isn't symmetric (e.g. male elves + female humans only).
    """
    vmap, bad = {}, []
    if cell is None:
        return vmap, bad
    for part in str(cell).split(","):
        p = part.strip()
        if not p or p.lower() == "none":
            continue
        if "_" not in p:
            bad.append(p)
            continue
        race, gender = p.rsplit("_", 1)
        race, gender = race.strip(), gender.strip()
        if not race or not gender:
            bad.append(p)
            continue
        vmap.setdefault(race, [])
        if gender not in vmap[race]:
            vmap[race].append(gender)
    return vmap, bad


def build_valid_map(races: list, genders: list, categories: dict) -> dict:
    """Per-race gender validity.

    The categories column wins when present (it can express asymmetric combos);
    otherwise fall back to the plain cross product of Valid Races x Valid
    Genders so rows without that column still work.
    """
    if categories:
        return {r: list(g) for r, g in categories.items() if g}
    return {r: list(genders) for r in races if genders}


def split_ids(cell) -> list:
    """Parse the Requirements cell: comma-separated requirement IDs."""
    if cell is None:
        return []
    out = []
    for part in str(cell).split(","):
        p = part.strip()
        if p and p.lower() != "none" and p not in out:
            out.append(p)
    return out


def split_cond_reqs(cell):
    """Parse the 'Race/Gender Specific Requirements' cell.

    Entries are "Condition:RequirementID", where Condition is either a race
    type ("Khajiit:KS_SMP_Khajiit") or a Race_Gender combo
    ("Khajiit_Female:Some_ID"). Commas separate entries. Returns
    ({condition: [ids]}, bad[]).
    """
    out, bad = {}, []
    if cell is None:
        return out, bad
    for part in str(cell).split(","):
        p = part.strip()
        if not p or p.lower() == "none":
            continue
        if ":" not in p:
            bad.append(p)
            continue
        cond, rid = p.split(":", 1)
        cond, rid = cond.strip(), rid.strip()
        if not cond or not rid:
            bad.append(p)
            continue
        out.setdefault(cond, [])
        if rid not in out[cond]:
            out[cond].append(rid)
    return out, bad


def cond_race(cond: str) -> str:
    """The race a condition applies to ("Khajiit" or "Khajiit_Female" -> Khajiit)."""
    if cond in KNOWN_RACE_TYPES:
        return cond
    if "_" in cond:
        base, last = cond.rsplit("_", 1)
        if last in KNOWN_GENDERS:
            return base
    return cond


def read_requirements(wb):
    """Read the requirements tab into an ordered list of {id, text, url}."""
    if REQ_SHEET not in wb.sheetnames:
        return [], [f"missing sheet {REQ_SHEET!r}"]
    ws = wb[REQ_SHEET]
    reqs, problems, seen = [], [], set()
    for r in range(2, ws.max_row + 1):
        rid = ws.cell(r, 1).value
        if rid is None or not str(rid).strip():
            continue
        rid = str(rid).strip()
        text = ws.cell(r, 2).value
        url = ws.cell(r, 3).value
        text = str(text).strip() if text else ""
        url = str(url).strip() if url else ""
        if rid in seen:
            problems.append(f"duplicate requirement ID {rid!r} (row {r})")
            continue
        seen.add(rid)
        if not text:
            problems.append(f"requirement {rid!r} has no link text")
        if not url:
            problems.append(f"requirement {rid!r} has no link")
        reqs.append({"id": rid, "text": text or rid, "url": url})
    return reqs, problems


def split_list(cell) -> list:
    """Split a comma-separated cell into a clean list, dropping blanks/None."""
    if cell is None:
        return []
    out = []
    for part in str(cell).split(","):
        p = part.strip()
        if p and p.lower() != "none":
            out.append(p)
    return out


def load_sheet(wb, name):
    if name not in wb.sheetnames:
        print(f"ERROR: sheet '{name}' not found. Sheets present: {wb.sheetnames}", file=sys.stderr)
        raise SystemExit(2)
    return wb[name]


def main():
    ap = argparse.ArgumentParser(description="Export spreadsheet tables to tool JSON.")
    ap.add_argument("--xlsx", required=True, help="Path to the spreadsheet")
    ap.add_argument("--previews", default="previews.json", help="Path to previews.json")
    ap.add_argument("--outdir", default="data", help="Directory to write JSON into (created if missing)")
    ap.add_argument("--report", default="export_report.txt", help="Validation report path")
    args = ap.parse_args()

    outdir = Path(args.outdir)
    outdir.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.load_workbook(args.xlsx, data_only=True)

    # ---- Race map (first two columns) ----
    ws = load_sheet(wb, "Race Default Hairstyles")
    race_map = {}
    for r in range(2, ws.max_row + 1):
        race = ws.cell(r, 1).value
        rtype = ws.cell(r, 2).value
        if race:
            race_map[str(race).strip()] = (str(rtype).strip() if rtype else None)

    # ---- Hairstyles ----
    ws = load_sheet(wb, "Hairstyles")
    hairstyles = []
    key_to_rows = defaultdict(list)   # detect collisions
    empty_valid = []
    bad_pairs = []                    # malformed entries in the exceptions column
    bad_cond_entries = []             # malformed entries in the conditional-req column
    asym = []                         # rows using per-race exceptions (for the report)
    hair_editorid_to_style = {}       # in-game hair EditorID -> hairstyle rec
    bald_style = None
    for r in range(2, ws.max_row + 1):
        mod = ws.cell(r, 1).value
        editor_ids = ws.cell(r, 2).value
        name = ws.cell(r, 3).value
        spell = ws.cell(r, 6).value
        if not (name and spell):
            continue
        name = str(name).strip()
        races = split_list(ws.cell(r, 7).value)
        genders = split_list(ws.cell(r, 8).value)
        categories, bad_cats = split_categories(ws.cell(r, 9).value)   # col 9
        for bad in bad_cats:
            bad_pairs.append((name, bad))
        reqs_ids = split_ids(ws.cell(r, 10).value)                     # col 10
        cond_reqs, bad_conds = split_cond_reqs(ws.cell(r, 11).value)   # col 11
        for bad in bad_conds:
            bad_cond_entries.append((name, bad))
        valid_map = build_valid_map(races, genders, categories)
        # Flat lists stay in sync with the map so older consumers still work.
        races = sorted(valid_map.keys())
        genders = sorted({g for gs in valid_map.values() for g in gs})
        key = preview_key(name)
        rec = {
            "mod": str(mod).strip() if mod else "",
            "name": name,
            "spell": str(spell).strip(),
            "previewKey": key,
            "validRaceTypes": races,
            "validGenders": genders,
            "validGendersByRace": valid_map,
            "requirements": reqs_ids,
            "condRequirements": cond_reqs,
            "editorIDs": split_list(editor_ids),
            "pack": pack_for(str(mod).strip() if mod else "", key),
            "displayName": NAME_OVERRIDES.get(name, name),
        }
        hairstyles.append(rec)
        key_to_rows[key].append(name)
        # Flag rows whose combos aren't the plain cross product -- these are the
        # asymmetric ones only the categories column can express.
        if valid_map and set(valid_map) and any(
                sorted(gs) != sorted(genders) for gs in valid_map.values()):
            asym.append((name, valid_map))
        if not races or not genders:
            empty_valid.append((name, races, genders))
        # index each in-game hair EditorID so NPC vanilla hair can be resolved
        for hid in split_list(editor_ids):
            hair_editorid_to_style.setdefault(hid, rec)
        if name.lower().endswith("bald") or key.lower().endswith("bald"):
            bald_style = rec

    def resolve_default(vanilla_hair):
        """Map an NPC's Vanilla Hair ID to a hairstyle rec (or None)."""
        if not vanilla_hair:
            return None
        vh = str(vanilla_hair).strip()
        if vh in hair_editorid_to_style:
            return hair_editorid_to_style[vh]
        if vh.lower() in ("(none/bald)", "none/bald", "bald", "none"):
            return bald_style
        return None

    # ---- NPCs ----
    ws = load_sheet(wb, "NPC List")
    npcs = []
    npc_race_missing = set()
    for r in range(2, ws.max_row + 1):
        eid = ws.cell(r, 2).value
        if not eid:
            continue
        race = str(ws.cell(r, 4).value).strip() if ws.cell(r, 4).value else ""
        rtype = race_map.get(race)
        if race and race not in race_map:
            npc_race_missing.add(race)
        vanilla_hair = str(ws.cell(r, 6).value).strip() if ws.cell(r, 6).value else ""
        default_style = resolve_default(vanilla_hair)
        npcs.append({
            "formid": str(ws.cell(r, 1).value).strip() if ws.cell(r, 1).value else "",
            "editorid": str(eid).strip(),
            "name": str(ws.cell(r, 3).value).strip() if ws.cell(r, 3).value else "",
            "race": race,
            "raceType": rtype,
            "gender": str(ws.cell(r, 5).value).strip() if ws.cell(r, 5).value else "",
            "vanillaHair": vanilla_hair,
            "defaultPreviewKey": default_style["previewKey"] if default_style else None,
            "defaultHairstyleName": default_style["name"] if default_style else None,
            "defaultSpell": default_style["spell"] if default_style else None,
        })

    # ---- Previews: build key -> {race: {gender: file}} ----
    previews = json.loads(Path(args.previews).read_text(encoding="utf-8"))
    pv_index = defaultdict(lambda: defaultdict(dict))
    pv_keys = set()
    for img in previews["images"]:
        pv_index[img["hairstyle"]][img["race"]][img["gender"]] = img["file"]
        pv_keys.add(img["hairstyle"])

    # ---- Validation ----
    hs_keys = {h["previewKey"] for h in hairstyles}
    no_preview = sorted(k for k in hs_keys if k not in pv_keys)
    orphan_imgs = sorted(pv_keys - hs_keys)
    collisions = {k: v for k, v in key_to_rows.items() if len(v) > 1}
    bad_race_types = sorted({t for h in hairstyles for t in h["validRaceTypes"]
                             if t not in KNOWN_RACE_TYPES})
    bad_genders = sorted({g for h in hairstyles for g in h["validGenders"]
                          if g not in KNOWN_GENDERS})

    # ---- Write JSON ----
    (outdir / "races.json").write_text(json.dumps(race_map, indent=2, ensure_ascii=True), encoding="utf-8")
    (outdir / "hairstyles.json").write_text(json.dumps(hairstyles, indent=2, ensure_ascii=True), encoding="utf-8")
    (outdir / "npcs.json").write_text(json.dumps(npcs, indent=2, ensure_ascii=True), encoding="utf-8")
    requirements, req_problems = read_requirements(wb)
    req_ids = {r["id"] for r in requirements}
    used_ids = Counter()
    undefined_refs = []      # (hairstyle, id)
    dead_conditions = []     # (hairstyle, condition, id, valid races)
    for h in hairstyles:
        for rid in h["requirements"]:
            used_ids[rid] += 1
            if rid not in req_ids:
                undefined_refs.append((h["name"], rid))
        valid_races = set(h["validGendersByRace"].keys())
        for cond, ids in h["condRequirements"].items():
            for rid in ids:
                used_ids[rid] += 1
                if rid not in req_ids:
                    undefined_refs.append((h["name"], rid))
                if cond_race(cond) not in valid_races:
                    dead_conditions.append((h["name"], cond, rid, sorted(valid_races)))
    unused_ids = sorted(req_ids - set(used_ids))
    no_reqs = [h["name"] for h in hairstyles if not h["requirements"]]
    (outdir / "requirements.json").write_text(
        json.dumps(requirements, indent=2, ensure_ascii=True), encoding="utf-8")

    # ---- Report ----
    lines = []
    lines.append("Export report")
    lines.append("=============")
    lines.append(f"NPCs exported          : {len(npcs)}")
    lines.append(f"Hairstyles exported    : {len(hairstyles)}")
    lines.append(f"Races in map           : {len(race_map)}")
    lines.append(f"Preview images         : {len(previews['images'])}  ({len(pv_keys)} unique keys)")
    lines.append("")
    lines.append(f"Hairstyles with NO preview image ({len(no_preview)}) -> tool shows a placeholder:")
    for k in no_preview:
        lines.append(f"   {k}   (from: {', '.join(key_to_rows[k])})")
    lines.append("")
    lines.append(f"Preview images with NO hairstyle row ({len(orphan_imgs)}):")
    for k in orphan_imgs:
        lines.append(f"   {k}")
    lines.append("")
    lines.append(f"Hairstyle rows with empty/None valid race or gender ({len(empty_valid)}):")
    for name, rc, ge in empty_valid:
        lines.append(f"   {name}   races={rc}  genders={ge}")
    lines.append("")
    lines.append(f"Preview-key collisions ({len(collisions)}):")
    for k, v in collisions.items():
        lines.append(f"   {k}: {v}")
    lines.append("")
    lines.append(f"Hairstyles with asymmetric race/gender combos ({len(asym)}):")
    for name, vmap in asym:
        lines.append(f"   {name}: " + "; ".join(f"{r}={'/'.join(g)}" for r, g in sorted(vmap.items())))
    lines.append("")
    lines.append(f"Malformed entries in the Race_Gender Categories column ({len(bad_pairs)}) -- expected Race_Gender:")
    for name, bad in bad_pairs:
        lines.append(f"   {name}: {bad!r}")
    lines.append("")
    lines.append(f"Requirements defined on the {REQ_SHEET!r} tab: {len(requirements)}")
    for p in req_problems:
        lines.append(f"   PROBLEM: {p}")
    lines.append("")
    lines.append(f"Requirement IDs referenced but NOT defined ({len(undefined_refs)}):")
    for name, rid in undefined_refs:
        lines.append(f"   {name}: {rid!r}")
    lines.append("")
    lines.append(f"Requirements defined but never used ({len(unused_ids)}): {unused_ids}")
    lines.append("")
    lines.append(f"Hairstyles with NO requirement at all ({len(no_reqs)}):")
    for n in no_reqs:
        lines.append(f"   {n}")
    lines.append("")
    lines.append(f"Conditional requirements for a race the hairstyle isn't valid for ({len(dead_conditions)}):")
    for name, cond, rid, vr in dead_conditions:
        lines.append(f"   {name}: {cond}:{rid}  (valid races: {vr})")
    lines.append("")
    lines.append(f"Malformed entries in the Race/Gender Specific Requirements column ({len(bad_cond_entries)}) -- expected Condition:RequirementID:")
    for name, bad in bad_cond_entries:
        lines.append(f"   {name}: {bad!r}")
    lines.append("")
    lines.append(f"Unknown gender values ({len(bad_genders)}): {bad_genders}")
    lines.append(f"Unknown valid-race-type values ({len(bad_race_types)}): {bad_race_types}")
    lines.append(f"NPC races missing from race map ({len(npc_race_missing)}): {sorted(npc_race_missing)}")
    report = "\n".join(lines)
    Path(args.report).write_text(report, encoding="utf-8")
    print(report)
    print(f"\nWrote {outdir/'npcs.json'}, {outdir/'hairstyles.json'}, {outdir/'races.json'} and {args.report}.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
